from openpyxl import load_workbook
from termcolor import colored

import base64
import json
import smtplib
import sys

def mail_serv(receiver,message):

    config_file = open('config.py','r')
    config = json.loads(config_file.read())
    config_file.close()
    mail_id     = config['EMAIL_ID']
    password    = config['PASSWORD']

    s = smtplib.SMTP('smtp.gmail.com', 587)

    s.starttls()
    s.login(mail_id, password)
    try:
        s.sendmail(mail_id, receiver, message) 
        print(colored(f'Mail sent to {receiver}','green'))
    except Exception:
        print(colored(f'Mail failed to send {receiver}','red'))
        pass

    s.quit()

def handle_xlsheet(path,msg):

    xl = load_workbook(path)
    ws = xl.active
    max_row = ws.max_row

    for row in range(1, max_row+1):
        email = ws['enter column of your exel sheet'+str(row)].value #mention Your Exel Coulnm Eg:- 'A' or 'B' like this
        mail_serv(email,msg)

        print(email)
        
if __name__=='__main__':

    if len(sys.argv) == 2:
        path = sys.argv[1]
        config_file = open('config.py','r')
        config = json.loads(config_file.read())
        config_file.close()

        message = """Subject: enter Your Syubject \n\n

                   Put Your Messsage Here

                  
                  """

        handle_xlsheet(path,message)
    else:
        print(f"{sys.argv[0]} <path to xlsx file>")
